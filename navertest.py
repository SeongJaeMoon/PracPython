import requests
from bs4 import BeautifulSoup
import io
import sys
import re
from openpyxl import Workbook, load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')


headers = {
    'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36'
}

cookies = {
'_naver_usersession_':'eP0GVZOV6WQSD7nsjUYNEQ=='
}

# def get_code(url):
#     req = requests.get(url, headers = headers, cookies = cookies)
#     return req.status_code

def get_data(url):
    try:
        req = requests.get(url, headers = headers, timeout = 5)
        bs = BeautifulSoup(req.content, 'html.parser')
        broad = ""
        title = ""
        contents = ""
        regdate = ""
        logo = bs.find('div', class_='press_logo')

        if logo is not None:
            broad = logo.select_one('img').attrs['title'] # 방송국
        if bs.find('h3', id = 'articleTitle') is not None:
            title = bs.find('h3', id = 'articleTitle').text # 기사 제목
        if  bs.find('div', id='articleBodyContents') is not None:
            contents = bs.find('div', id='articleBodyContents').text.replace('// flash 오류를 우회하기 위한 함수 추가', '').replace('function _flash_removeCallback() {}', '') # 기사 내용
        if bs.find('span', class_='t11') is not None:
            regdate = bs.find('span', class_='t11').text # 등록일
        return broad, title, contents, regdate
    except:
        return None

def get_links(links):
    link_ret =[]
    for a in links:
        href = a.attrs['href'] # 주소
        if href is not '#':
            link_ret.append(href)
    return link_ret

url = "https://search.naver.com/search.naver?&where=news&query=%EC%84%9C%EC%9A%B8%20%EC%8B%9C%EC%9E%A5&sm=tab_pge&sort=0&photo=0&field=0&reporter_article=&pd=3&ds=2014.05.01&de=2014.05.31&docid=&nso=so:r,p:from20140501to20140531,a:all&mynews=0&start={}&refresh_start=0"
url2 = "https://search.naver.com/search.naver?&where=news&query=%EB%B0%95%EC%9B%90%EC%88%9C&sm=tab_pge&sort=0&photo=0&field=0&reporter_article=&pd=3&ds=2014.05.01&de=2014.05.31&docid=&nso=so:r,p:from20140501to20140531,a:all&mynews=0&start={}&refresh_start=0"
url3 = "https://search.naver.com/search.naver?&where=news&query=%EC%A0%95%EB%AA%BD%EC%A4%80&sm=tab_pge&sort=0&photo=0&field=0&reporter_article=&pd=3&ds=2014.05.01&de=2014.05.31&docid=&nso=so:r,p:from20140501to20140531,a:all&mynews=0&start={}&refresh_start=0"

try:
    wb = load_workbook('C:\\Users\\SIST\\Documents\\moonsworld\\naver_news.xlsx')
    ws = wb.create_sheet(title='')
    ws['A1'] = '보도국'
    ws['B1'] = '제목'
    ws['C1'] = '기사내용'
    ws['D1'] = '등록일'
    total = 1
    for page in range(0, 40):
        idx = (page * 10) + 1
        req = requests.get(url3.format(str(idx)), headers = headers, timeout = 5)
        # print(req.status_code)
        bs = BeautifulSoup(req.content, 'html.parser')

        total_list = bs.find('ul', class_ = 'type01')
        links = total_list.select("a[href]") # url
        # links = total_list.select('dl')

        link_ret = get_links(links)
        # print(link_ret)
        regex = r'[가-힣]+'
        p = re.compile(regex)

        for page in link_ret:
            if get_data(page) is None:
                print('404 pages')
            else:
                broad, title, contents, regdate = get_data(page)
                m2 = p.search("서울시장")
                m3 = p.search("박원순")
                m4 = p.search("정몽준")

                if m2 is not None or m3 is not None or m4 is not None:
                    if 'SBS' in broad or 'KBS' in broad or 'MBC' in broad:
                        try:
                            # print(broad, title, contents, regdate)
                            ws['A'+str(1 + total)] = broad
                            ws['B'+str(1 + total)] = title
                            ws['C'+str(1 + total)] = contents
                            ws['D'+str(1 + total)] = regdate
                            total += 1
                            wb.save('C:\\Users\\SIST\\Documents\\moonsworld\\naver_news.xlsx')
                        except:
                            print('Save Error!!')
                        finally:
                            wb.close()
finally:
    wb.close()
