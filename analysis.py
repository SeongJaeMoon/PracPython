from openpyxl import load_workbook
from konlpy.tag import Twitter
import matplotlib as mpl
from matplotlib import pyplot as plt
import matplotlib.font_manager as fm
import pandas as pd
import numpy as np
import re
import string

# 그래프에서 한끌 깨짐 방지
mpl.rcParams['axes.unicode_minus'] = False
path = '/Library/Fonts/NanumGothic.ttf'
font_name = fm.FontProperties(fname=path, size=10)

# 파일이름 : 1993 한국증권학회 : 심포지엄 Sheet / C 발표주제 > 1993_ko.xlsx, rows = 3:430
# 파일이름 : 재무관리학회_학회지(완료) : 재무관리연구 Sheet / C 논문제목 > jaemoo.xlsx
# 파일이름 : 파생상품학회_학회지외 (1)(완료) : 파생상품학회_심포지엄 Sheet / C 논문제목 > product.xlsx - 파생상품학회_선물연구 Sheet로
# 파일이름 : 재무학회 현황(재무연구, 심포지엄, 학술포럼) : 재무학회-심포지엄 Sheet / c 발표주제 > money.xlsx
# 파일이름 : 공동학술대회 자료 : 심포지엄과 발표논문 2개의 sheet 모두/ 심포지엄 sheet는 c 발표주제 + 발표논문 sheet는 b 논문제목 > coacademic.xlsx
'''
파생상품학회_학회지외 파일의 파생상품학회_심포지엄 시트에 논문제목 컬럼이 존재하지 않아 파생상품학회_선물연구에서 논문제목 컬럼을 추출하여 분석했습니다.
'''
dict_data = {}
tw = Twitter()
FILE_PATH = "/Users/moonseongjae/Downloads/coacdemic.xlsx"
SAVE_PATH = "/Users/moonseongjae/python-worksapce/Project/crawling_test/result.xlsx"
try:
    wb = load_workbook(FILE_PATH)
    content_list = []
    sheet = wb['공동학술대회-발표논문'] # 공동학술대회-발표논문
    for i in range(3, 1341): # 1341
        content = sheet["B" + str(i)].value
        if content is not None:
            content_list.append(content)
            # print(content)

  
    for alt in content_list:
        match_pattern = re.findall('[a-zA-Z]{3,15}', alt)
        temp = tw.pos(alt, norm = True)
        for data in temp:
            if data[1] == "Noun":
                if not (data[0] in dict_data):
                    dict_data[data[0]] = 0
                dict_data[data[0]] += 1
        for en in match_pattern:
            if not (en in dict_data) and (en.isalpha()):
                dict_data[en] = 0
            dict_data[en] += 1
    
    keys = sorted(dict_data.items(), key = lambda x:x[1], reverse = True)
    wb = load_workbook(SAVE_PATH)
    sub_name =[]
    vals = []
    idx = 0
    total = 2
    ws = wb.active
    for k, v in keys:
        # print("{}({})".format(k, v))
        ret = "{}({})".format(k, v)
        ws["F" + str(total)] = ret
        if idx < 20: 
            sub_name.append(k)
            vals.append(v)
        idx += 1
        total += 1
        wb.save(SAVE_PATH)

    y_pos = np.arange(20)
    plt.rcdefaults()
    fig, ax = plt.subplots()
    ax.barh(y_pos, vals, align='center', color='green', ecolor='black')
    ax.set_yticks(y_pos)
    ax.set_yticklabels(sub_name, fontproperties=font_name)
    ax.invert_yaxis()  # labels read top-to-bottom
    ax.set_xlabel('횟수', fontproperties=font_name)
    ax.set_title('자주 등장한 상위 20개 명사?', fontproperties=font_name)
    plt.savefig("공동학술대회_발표논문.jpg")
    plt.show()    
finally:
    wb.close()
    
