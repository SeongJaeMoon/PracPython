import sys
import io
import json
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

# 2016.10 ~ 2018.03 (월 한 번 크롤링(1~1000개) 하는데 5분 이상 소요... 많을 때는 10분 정도 소요.)
# 달 단위 -> sheet 총 18 sheets -> 18 * 1000 = 18000
# 큰 단락 -> 월간 순위
# ddate -> 월 : 2016-10 ~ 2018-03, start -> 시작 페이지 : 1, 101, 201, ..., 901. , , 1, 2, ... 9

# 순위(자동 증가 = total), 상승 or 하락 순위, 카테고리, 제목, 부가설명, UP날짜, 청취자수, 시청자수, 좋아요수, 댓글수, 구독수. -> 함수로 정의 필요.

# 상승, 하락 순위 함수.
def get_up_down(td_list):
    up_downs = ''
    value = ''
    if 'rank_up' in td_list[1] or 'rank_down' in td_list[1]:
        # 상승, 하락 알기 위한 임시 변수.
        # up_down = str(td_list[1]).split('.png"/>')[2].split('icon_rank_')[1]
        up_down = str(td_list[1]).split('icon_rank_')[1].split('.png')[0]
        up_downs = ''
        if 'down' in up_down:
            up_downs = '하락'
        else :
            up_downs = '상승'
        # 상승 or 하락 순위.
        if len(td_list[1].split('.png"/>')) == 6:
            value = str(td_list[1]).split('.png"/>')[4].split('</p>')[0]
        elif len(td_list[1].split('.png"/>')) < 5:
            value = str(td_list[1]).split('.png"/>')[2].split('</p>')[0]
        else:
            value = str(td_list[1]).split('.png"/>')[3].split('</p>')[0]

    return up_downs, value


# 카테고리 함수.
def get_category(td_list):
    cate = td_list[3].split('section_2">')[1].split('/category/')[1].split('">')[1].split('</a>')[0]
    # print(td_list[3].split('section_2">')[1].split('/category/')[1].split('">')[1].split('</a>')[0])
    return cate

# 제목, 부가설명, UP날짜, 청취자수, 시청자수 함수.
def get_inner(td_list):
    title = td_list[3].split('title="')[1].split('">')[0] # 제목.
    explain = str(td_list).split('[카테고리]')[1].split('</dd>')[2].split('>')[1].replace('\\n', '').replace('\\t','')
    update = td_list[3].split('<li>')[2].split('/>')[1].replace('</li>', '') # UP날짜.
    listen = td_list[3].split('<li>')[3].split('/>')[1].replace('</li>','') # 청취자수.
    viwer = td_list[3].split('<li>')[4].split('icon_play.png"/>')[1].split('</li>')[0] # 시청자수.

    return title, explain, update, listen, viwer

# 좋아요수, 댓글수, 구독수.
def get_number(td_list):
    star = td_list[3].split('star_number">')[1].split('</p>')[0] # 좋아요 수.
    comment = td_list[3].split('etc_number')[1].split('</p>')[0].replace('">', '') # 댓글 수.
    subscribe = td_list[3].split('etc_number')[2].split('</p>')[0].replace('">', '') # 구독 수.

    return star, comment, subscribe

try:
        wb = Workbook()
#     year = 2015
#     month = 2
#     pages = 100
# # 순위(자동 증가 = total), 상승 or 하락 순위, 카테고리, 제목, 부가설명, UP날짜, 청취자수, 시청자수, 좋아요수, 댓글수, 구독수.
#     for i in range(3): # 년도만큼 반복.
#         year += 1
        # month += 1
        # while True: # 월만큼 반복.
        #     month = 3
        wb = load_workbook('C:\\Users\\SIST\\Documents\\moonsworld\\pb2.xlsx')
        ws = wb.create_sheet(title = '2016-04') # 시트 이름 바꿔줘야함.
        ws['A1'] = '순위'
        ws['B1'] = '상승or하락'
        ws['C1'] = '카테고리'
        ws['D1'] = '제목'
        ws['E1'] = '부가설명'
        ws['F1'] = 'UP날짜'
        ws['G1'] = '청취자수'
        ws['H1'] = '시청자수'
        ws['I1'] = '좋아요수'
        ws['J1'] = '댓글수'
        ws['K1'] = '구독수'
        total = 1 # 자동 증가 변수.(컬럼 증가.)
        rank = 1 # 순위.
        for page in range(0, 10): # 1 ~ 901까지 반복.
            url = 'http://www.podbbang.com/ranking?kind=monthly&ddate=2016'+'-'+'{}&start={}'
            req = requests.get(url.format(str(4), str((100 * page) + 1))) #url 년도, 월 바꿔줘야함.

            bs = BeautifulSoup(req.content.decode('utf-8', 'replace'), 'html.parser')
            #테이블 태그.
            table = bs.find('table', class_ = 'tb_rank')

            # 모든 tr_tag 가져오기. - 하나의 행.
            tr = table.find_all('tr')
            # 모든 td-tag 가져오기. 하나의 열.
            print('len:',len(tr))

            for idx in range(0, len(tr)): #tr 크기만큼 반복.
                td_list = str(tr[idx]).split('td')

                ws['A'+str(1 + total)] = rank
                up_downs, value = get_up_down(td_list)
                ws['B'+str(1 + total)] = up_downs + str(value)
                ws['C'+str(1 + total)] = get_category(td_list)
                title, explain, update, listen, viwer = get_inner(td_list)
                ws['D'+str(1 + total)] = title
                ws['E'+str(1 + total)] = explain
                ws['F'+str(1 + total)] = update
                ws['G'+str(1 + total)] = listen
                ws['H'+str(1 + total)] = viwer
                star, comment, subscribe = get_number(td_list)
                ws['I'+str(1 + total)] = star
                ws['J'+str(1 + total)] = comment
                ws['K'+str(1 + total)] = subscribe
                total += 1

                wb.save('C:\\Users\\SIST\\Documents\\moonsworld\\pb2.xlsx')
                if idx == (len(tr)-1):
                    # 1~991
                    rank = (100 * (page + 1)) + 1
                    print('rank:', rank)
                else:
                    rank += 1
            #     month += 1 # 월 하나 증가시키기.
            # if month == 13:
            #     month = 0
            #     break
finally:
    wb.close()
