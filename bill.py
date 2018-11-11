from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import time
import requests
from bs4 import BeautifulSoup
import json
from openpyxl import load_workbook

# 검색기간 : 2018.05 ~ 2018.08
# 검색어 : 정보통신, 전기통신, 방송통신, 인터넷, 가짜, 디지털, 글로벌, 위치, 공간, 4차, 저작권, 보도, 선거, 조세, 개인정보보호(법률안만, 해당, 선출안 제외)
# 기본 -> (의안번호), 의안명, 제안자 구분, 제안자(세부)<대표 발의자 처음 표시, 정당 함께 표시>, 제안일자, 의결일자, 의결결과, 심사진행상태
# 상세 -> 제안이유 + 주요내용(세부), 제안회기(세부), 소위원회(세부), 관련위원회(세부)

BASE_URL = "http://likms.assembly.go.kr/bill/main.do" # 기본 URL
DRIVER_DIR = '/Users/moonseongjae/chromedriver' # 크롬 드라이버 경로
DETAIL_URL = "http://likms.assembly.go.kr/bill/billDetail.do?billId=" # 상세 경로
USER_URL = "http://likms.assembly.go.kr/bill/coactorListPopup.do?billId=" # 발의 의원 정보 경로
SAVE_DIR = '/Users/moonseongjae/Proejct_FC/temp/temp.xlsx' # 엑셀 저장 경로

HEADER = {
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 Edge/16.16299',
    'referer':'rehttp://likms.assembly.go.kr/bill/main.do'
}

# Proposal
def get_scrap(scrap):
    driver = webdriver.Chrome(DRIVER_DIR)
    wb = load_workbook(SAVE_DIR)
    try:
        for scr in scrap:
            print(scr)
            ws = wb.create_sheet(title = str(scr))
           
            # ws['A1'] = '의안번호'
            # ws['B1'] = '의안명'
            # ws['C1'] = '제안자구분'
            # ws['D1'] = '제안자(세부)'
            # ws['E1'] = '제안일자'
            # ws['F1'] = '의결일자'
            # ws['G1'] = '의결결과'
            # ws['H1'] = '심사진행상태'
            # ws['I1'] = '주요내용'
            # ws['J1'] = '제안회기'
            # ws['K1'] = '소관위원회'
            # ws['L1'] = '관련위원회'

            driver.get(BASE_URL)
            e = driver.find_element_by_xpath('//*[@id="si1_label05"]') # 의안명 입력창
            e.clear()
            e.send_keys(str(scr)) # 의안명 입력
            e.send_keys(Keys.ENTER) # 엔터
            sleep(1)
            page_size = driver.find_element_by_xpath('//*[@id="pageSizeOption"]') # 페이지당 결과 수
            page_size.click()
            sleep(1)
            page100 = driver.find_element_by_xpath('//*[@id="pageSizeOption"]/option[4]') # 페이지당 결과 수 100으로 변경
            page100.click() 
            sleep(1)
            second = driver.find_element_by_xpath('//*[@id="pageListViewArea"]/a[2]')
            second.click()
            sleep(1)
            tbody = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[2]/table/tbody')

            result_list = [] # 결과 리스트
            detail_list = [] # 상세 페이지 리스트
            
            for tr in tbody.find_elements_by_tag_name('tr'):                
                data = {}
                td = tr.find_elements_by_tag_name('td')
                if td is not None:
                    prop_date = td[3].text # 제안 일자
                    isDate = str(prop_date).split('-')[1] # 구분자 
                    
                    if int(isDate) < 5: break # 제안 일자가 5(5월) 보다 작으면 반복 정지

                    temp = td[1].find_element_by_css_selector('div.pl25 > a') 
                    link = str(temp.get_attribute('href')).split("fGoDetail('")[1].split("',")[0] # 상세 링크
                
                    detail_list.append(link) 

                    data['number'] = td[0].text # 의안 번호 
                    data['title'] = temp.get_attribute('title') # 의안명
                    data['proponent'] = td[2].text # 제안자 구분
                    data['propdate'] = prop_date # 제안 일자
                    data['decidate'] = td[4].text # 의결 일자
                    data['deciret'] = td[5].text # 의결 결과
                    data['status'] = td[7].text # 심사진행상태
                    result_list.append(data)
            
            detail_result_list = [] # 상세 정보
            for link in detail_list:
                data = {}
                driver.get(DETAIL_URL + link)
                sleep(1)
                members = []
                try:
                    e = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]/a/img')
                    req = requests.get(USER_URL + link)
                    soup = BeautifulSoup(req.text, 'html.parser')
                    for a in soup.select('div.mt20 a'):
                        members.append(a.get_text())
                except Exception as e:
                    member = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[3]').text
                    members.append(member)

                sugg = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[3]/div[1]/table/tbody/tr/td[5]').text # 제안회기
                
                content = None
                try:
                    content = driver.find_element_by_id('summaryContentDiv').text # 주요 내용
                except Exception as e:
                    content = ''

                committee = None
                try:
                    committee = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[5]/div/table/tbody/tr/td[1]').text # 소관위원회
                except Exception as e:
                    committee = ''
                relation = None
                try:
                    relation = driver.find_element_by_xpath('/html/body/div/div[2]/div[2]/div/div[5]/div[3]/table/tbody/tr/td[1]').text
                except Exception as e:
                    relation = ''
                    
                data['content'] = content # 주요내용
                data['sugg'] = sugg # 제안회기
                data['committee'] = committee # 소관위
                data['relation'] = relation #관련위
                data['member'] = str(members).replace('[', '').replace(']', '') #발의 위원                
                detail_result_list.append(data)

            total = 0 
            for ret in result_list: # 기본
                ws['A'+str(2 + total)] = ret['number']
                ws['B'+str(2 + total)] = ret['title']
                ws['C'+str(2 + total)] = ret['proponent']
                ws['E'+str(2 + total)] = ret['propdate']
                ws['F'+str(2 + total)] = ret['decidate']
                ws['G'+str(2 + total)] = ret['deciret']
                ws['H'+str(2 + total)] = ret['status']
                total += 1

            total = 0
            for ret in detail_result_list: # 상세
                ws['D'+str(2 + total)] = ret['member']
                ws['I'+str(2 + total)] = ret['content']
                ws['J'+str(2 + total)] = ret['sugg']
                ws['K'+str(2 + total)] = ret['committee']
                ws['L'+str(2 + total)] = ret['relation']
                total += 1
            wb.save(SAVE_DIR)
    except Exception as e:
        print(e)
    finally:
        driver.close()
        wb.close()

if __name__ == "__main__":
    # search = ['정보통신']
    start_time = time.time()
    # '정보통신', '전기통신', '방송통신', '인터넷', '가짜'
    # 디지털은 -> '2017-03-07' 밖에 검색 결과가 없음, 글로벌 -> 검색 결과가 1건도 없음, 보도 -> 검색 결과가 1건도 없음
    search = ['조세']
    get_scrap(search)
    print("--- %s seconds ---" % (time.time() - start_time))
 





