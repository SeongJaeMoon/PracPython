
import sys
import io
import requests
import json
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')
# 정보통신, 방송통신, 전기통신, 인터넷, 정보보호(법률안만 해당), 가짜, 디지털
# (의안번호), 의안명, 제안자 구분, 제안자(세부)<대표 발의자 처음 표시, 정당 함께 표시>, 제안일자, 의결일자, 의결결과, 심사진행상태, 제안이유 + 주요내용(세부), 제안회기(세부), 소위원화(세부), 관련위원회(세부)
info_list = []
broad_list = []
elec_list = []
secure_list = []
fake_list = []
digital_list = []
# 개발자 도구 network 탭에서 URL 분석 필요.
# 정보통신 89건 - 9pages
url1 = "http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EC%A0%95%EB%B3%B4%ED%86%B5%EC%8B%A0&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage={}&pageSize=10&maxPage=10&billparam=&allCount=89&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch"

for i in range(1, 10):
    info_list.append(url1.format(str(i)))

# 방송통신 37건 - 4pages
url2 = "http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EB%B0%A9%EC%86%A1%ED%86%B5%EC%8B%A0&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage={}&pageSize=10&maxPage=10&billparam=&allCount=89&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch"

for i in range(1, 5):
    broad_list.append(url2.format(str(i)))

# 전기통신 63건 - 7pages
url3 ="http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EC%A0%84%EA%B8%B0%ED%86%B5%EC%8B%A0&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage={}&pageSize=10&maxPage=10&billparam=&allCount=37&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch"

for i in range(1, 8):
    elec_list.append(url3.format(str(i)))

# 인터넷 7건 - 1pages
url4= "http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EC%9D%B8%ED%84%B0%EB%84%B7&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage=1&pageSize=10&maxPage=10&billparam=&allCount=63&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch"

# 개인정보보호 총 24건 - 3pages
url5="http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EA%B0%9C%EC%9D%B8%EC%A0%95%EB%B3%B4%EB%B3%B4%ED%98%B8&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=%EC%A0%84%EC%B2%B4&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=%EC%A0%84%EC%B2%B4&generalResultType=&committeeResult=%EC%A0%84%EC%B2%B4&periodFrom=&periodTo=&budgetYn=%EC%A0%84%EC%B2%B4&billGbnAmendment=N&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage={}&pageSize=10&maxPage=10&billparam=&allCount=29&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billDetailSearch"

for i in range(1, 4):
    secure_list.append(url5.format(str(i)))

# 디지털 총 1건
url6="http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EB%94%94%EC%A7%80%ED%84%B8&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage=1&pageSize=10&maxPage=10&billparam=&allCount=102&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch"


# 가짜 총 1건
url7 = 'http://likms.assembly.go.kr/bill/BillSearchResult.do?billNo=&billName=%EA%B0%80%EC%A7%9C&simsaBonmun=&ageFrom=20&ageTo=20&age=&dateFrom=&dateTo=&procFrom=&procTo=&dispFrom=&dispTo=&proposeFrom=&proposeTo=&submitFrom=&submitTo=&billNoFrom=&billNoTo=&billKind=%EC%A0%84%EC%B2%B4&billKind2=&billKindExclude=&chargeCommittee=&chargeCommitteeName=&proposerKind=%EC%A0%84%EC%B2%B4&proposer=&hjNm=&statsType=&proposeGubn=%EC%A0%84%EC%B2%B4&generalResult=&generalResultType=&committeeResult=&periodFrom=&periodTo=&budgetYn=&billGbnAmendment=&orderColumn=Propose_Dt&orderType=DESC&lastQuery=&strPage=1&pageSize=10&maxPage=10&billparam=&allCount=89&resultCount=10&mooring=N&billNoYn=N&statsKind=&ageCheck=&tabMenuType=billSimpleSearch'

# 가짜 검색 상세보기 테스트 url
# http://likms.assembly.go.kr/bill/billDetail.do?billId=PRC_U1Y8C0J4Y0D5N1R7B5F6G2B2J0C9H8

# headers, cookies get 방식 요청. config (params = GET, data = POST)
headers = {
'Content-Type': 'application/x-www-form-urlencoded',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 Edge/16.16299',
'referer':'rehttp://likms.assembly.go.kr/bill/main.do'
}
cookies = {
'PHAROSVISITOR':'000014610162bfb72e1b4c210ac93c0d;',
'JSESSIONID':'hgK56VHq7qq5A7xt2aSGJUZKEsr9jHhGsQRy0p7xHihaubrjMainRAkt8TQ7pta1.amV1c19kb21haW4vYmlsbDE='
}
# 의안명 구하는 함수.
def get_det_url(tag):
    temp = tag.find('a').get('href')
    # detail_url
    det_url = temp.replace("', 'billSimpleSearch')",'').replace("javascript:fGoDetail('",'')
    return det_url


# 주요 내용 구하는 함수.
def get_detail_content(det_url):
    detail = requests.get('http://likms.assembly.go.kr/bill/billDetail.do?billId='+det_url)
    res = BeautifulSoup(detail.text, 'html.parser')
    table = res.find_all('div', class_='tableCol01')
    tr = res.find_all('tr')
    temp = res.find('div', id = 'summaryContentDiv')
    if temp is not None:
        contentData = temp.text
    else:
        contentData = ''
    # 제안 회기
    data7 = str(tr).split('<td>')[5].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','').split('>')[1].replace('</tr' , '')
    # print(data7)
    # 소관 위원회.
    data8 = ''
    if len(table) > 2:
        data8 = str(tr).split('<td>')[6].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','').split('>')[1].replace('</tr' , '')
    data9 = ''
    if len(table) > 3:
        data9 = str(tr).split('<td>')[15].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','').split('>')[1].replace('</tr' , '')
    return contentData, data7, data8, data9

# 의안 발의 국회의원 명단 구하는 함수.
def get_detail_list(det_url):
    saver = requests.get("http://likms.assembly.go.kr/bill/coactorListPopup.do?billId="+det_url)
    soup = BeautifulSoup(saver.text, 'html.parser')
    admin = soup.find('div', class_ ='links')
    # 의원 명단.
    coactor_list = []
    admin_val = admin.find_all('a')
    l = len(str(admin_val).split('>'))
    for i in range(1, l, 2):
        coactor_list.append(str(admin_val).split('>')[i].replace('</a', ''))
    return coactor_list

try:
    wb = Workbook()
    wb = load_workbook('/Users/moonseongjae/testinfo.xlsx')
    ws = wb.create_sheet(title='개인정보보호')
    # ws = wb.active
    # ws.title = '정보통신'
    ws['A1'] = '의안번호'
    ws['B1'] = '의안명'
    ws['C1'] = '제안자구분'
    ws['D1'] = '제안자(세부)'
    ws['E1'] = '제안일자'
    ws['F1'] = '의결일자'
    ws['G1'] = '의결결과'
    ws['H1'] = '심사진행상태'
    ws['I1'] = '주요내용'
    ws['J1'] = '제안회기'
    ws['K1'] = '소관위원회'
    ws['L1'] = '관련위원회'
    total = 0
    for idx in range(len(secure_list)):
        req = requests.get(secure_list[idx], headers = headers, cookies = cookies)
        soup = BeautifulSoup(req.text, 'html.parser')
        tbody = soup.find('tbody')
        tr = tbody.find_all('tr')
    # print(tr)

        for i in range(len(tr)):
            div_tag = tr[i].find('div', class_='pl25')
            temp = div_tag.find('a').text.replace(r'n|t|\\', '')
            # print(temp)
            if '선출안' not in temp:
                # detail url.
                det = get_det_url(div_tag).replace("', 'billDetailSearch')", '')
                # 의안번호.
                data0 = str(tr[i]).split('</td')[0].replace('<td>','').replace('<tr>', '')
                # 의안명.
                data1 = div_tag.find('a').text.replace(r'n|t|\\', '')
                # 제안자구분.
                data2 = str(tr[i]).split('<td>')[2].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','')
                # 제안일자.
                data3 = str(tr[i]).split('<td>')[3].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','')
                # 의견 일자.
                data4 = str(tr[i]).split('<td>')[4].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','')
                # 의견 결과.
                data5 = str(tr[i]).split('<td>')[5].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','')
                #심사 진행 상태.
                data6 = str(tr[i]).split('<td>')[7].replace('<tr>|</tr>|\n|</td>', '').replace('</td>','').replace('</tr>', '')
                # 세부 내용, 제안회기, 소관 위원회, 관련 위원회
                content, data7, data8, data9 = get_detail_content(det)
                # 제안자 목록
                coactor_list = get_detail_list(det)
                # 의안명, 제안자 구분, 제안자(세부)<대표 발의자 처음 표시, 정당 함께 표시>, 제안일자, 의결일자, 의결결과, 심사진행상태, 제안이유 + 주요내용(세부), 제안회기(세부), 소위원회(세부), 관련위원회(세부)

                # 데이터 저장.

                ws['A'+str(2 + total)] = data0
                ws['B'+str(2 + total)] = data1
                ws['C'+str(2 + total)] = data2
                ws['D'+str(2 + total)] = str(coactor_list)
                ws['E'+str(2 + total)] = data3
                ws['F'+str(2 + total)] = data4
                ws['G'+str(2 + total)] = data5
                ws['H'+str(2 + total)] = data6
                ws['I'+str(2 + total)] = content
                ws['J'+str(2 + total)] = data7
                ws['K'+str(2 + total)] = data8
                ws['L'+str(2 + total)] = data9
                total += 1
            # 엑셀 파일로 변환.
    wb.save('/Users/moonseongjae/testinfo.xlsx')
finally:
    wb.close()
